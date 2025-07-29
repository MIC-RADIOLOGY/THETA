import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
import uuid
import streamlit.components.v1 as components

st.set_page_config(page_title="Aging Report Merger", layout="wide")
st.title("📊 Aging Report Merger (ZWL & USD with Formulas)")

# --- Cleaning ---
def clean_dataframe(df):
    df = df.applymap(lambda x: str(x).replace(',', '') if isinstance(x, str) else x)
    df = df.apply(pd.to_numeric, errors='ignore')
    df = df.applymap(lambda x: 0 if isinstance(x, (int, float)) and x < 0 else x)
    return df

# --- Header detection ---
def read_excel_properly(file):
    raw = pd.read_excel(file, header=None)
    st.write("🔍 Raw Data Preview (first 10 rows):")
    st.dataframe(raw.head(10))

    keywords = ['customer', 'name', 'account', 'client', 'acct', 'provider']

    for i in range(len(raw)):
        row = raw.iloc[i]
        if row.count() < 2:
            continue
        valid = row.dropna().astype(str).tolist()
        if any(any(kw in cell.strip().lower() for kw in keywords) for cell in valid):
            df = pd.read_excel(file, skiprows=i)
            df.columns = df.columns.astype(str)
            return df
    raise ValueError("Could not detect valid header row. Please check the format of your file.")

# --- Excel writer ---
def to_excel_with_formulas(df, sample_file):
    sample_wb = load_workbook(sample_file, data_only=False)
    sample_ws = sample_wb.active

    formula_columns = {}
    for col_idx, cell in enumerate(sample_ws[2], start=1):
        if cell.data_type == 'f':
            formula_columns[col_idx - 1] = cell.value

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())

    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        for col_idx, formula in formula_columns.items():
            adjusted_formula = formula.replace("2", str(row_idx))
            ws.cell(row=row_idx, column=col_idx + 1).value = f"={adjusted_formula}"

    wb.save(output)
    output.seek(0)
    return output

# --- Upload section ---
st.header("📂 Step 1: Upload Files")
sample_file = st.file_uploader("📄 Upload Sample Layout File (with Formulas)", type=["xlsx"])
zwl_file = st.file_uploader("💱 Upload ZWL Aging Report", type=["xlsx"])
usd_file = st.file_uploader("💲 Upload USD Aging Report", type=["xlsx"])

# --- Processing ---
if not sample_file:
    st.warning("Please upload the sample layout file.")
elif not zwl_file:
    st.warning("Please upload the ZWL report.")
elif not usd_file:
    st.warning("Please upload the USD report.")
else:
    try:
        # Read sample layout headers
        sample_wb = load_workbook(sample_file, data_only=False)
        sample_ws = sample_wb.active
        sample_headers = [cell.value for cell in sample_ws[1]]

        # Read and clean ZWL
        zwl_df = read_excel_properly(zwl_file)
        zwl_df = clean_dataframe(zwl_df)
        zwl_df.insert(0, "Currency", "ZWL")
        if 'Balance' in zwl_df.columns:
            zwl_df.drop(columns=['Balance'], inplace=True)

        # Read and clean USD
        usd_df = read_excel_properly(usd_file)
        usd_df = clean_dataframe(usd_df)
        usd_df.insert(0, "Currency", "USD")
        if 'Balance' in usd_df.columns:
            usd_df.drop(columns=['Balance'], inplace=True)

        # Merge and align
        merged_df = pd.concat([zwl_df, usd_df], ignore_index=True)
        merged_df = merged_df.reindex(columns=sample_headers)

        st.subheader("✅ Step 2: Consolidated Preview")
        st.dataframe(merged_df)

        final_excel = to_excel_with_formulas(merged_df, sample_file)

        st.subheader("📥 Step 3: Download Final Report")
        # Fallback download button
        st.download_button(
            label="Download Consolidated Aging Report",
            data=final_excel,
            file_name="Consolidated_Aging_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Auto-download logic
        if "auto_dl_done" not in st.session_state or not st.session_state["auto_dl_done"]:
            file_bytes = final_excel.getvalue()
            b64 = base64.b64encode(file_bytes).decode()
            fname = "Consolidated_Aging_Report.xlsx"
            uid = str(uuid.uuid4()).replace("-", "")

            html = f"""
            <a id="dl-{uid}"
               href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
               download="{fname}" style="display:none">download</a>
            <script>
              const a = document.getElementById("dl-{uid}");
              if (a) a.click();
            </script>
            """
            components.html(html, height=0)
            st.session_state["auto_dl_done"] = True
            st.success("✅ Report ready. Your download should start automatically.")

    except Exception as e:
        st.error(f"❌ Error: {e}")
            
